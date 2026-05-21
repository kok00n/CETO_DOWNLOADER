"""Refresh bond_auctions from MF Obligacje_Hurtowe.xlsm.

Parses Operacje sheet, filters to auction-like operations
(TypOperacji in AS/AU/AZ/AO), extracts auction details per row.

Each row in MF Operacje is one leg of an auction event - e.g. switch
auction (AZ) has separate rows for the buyback leg (TypTransakcji='O')
and the sale leg (TypTransakcji='S').  We store both as bond_auctions
rows; downstream views (v_recent_auctions) filter to type_tx='S' for
sale-side metrics.

Idempotent via upsert on auction_id (= MF #ID, unique per row).
"""

from __future__ import annotations

import sys
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from lib.mf_xlsm import download_xlsm, find_xlsm_url  # noqa: E402
from lib.supabase import upsert  # noqa: E402


AUCTION_TYPE_OPS = {"AS", "AU", "AZ", "AO"}


def _to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _yield_to_pct(value) -> float | None:
    """MF stores yields as decimal (0.054 = 5.4%). Normalize to percent
    so they match BondSpot fixing_yield convention (3.56 = 3.56%)."""
    f = _to_float(value)
    return f * 100 if f is not None else None


def _to_id(value) -> str | None:
    """MF #ID stored as float-string; normalize to plain integer string."""
    if value is None or value == "":
        return None
    if isinstance(value, str):
        s = value.strip()
        if s.endswith(".0"):
            return s[:-2]
        return s
    try:
        return str(int(float(value)))
    except (TypeError, ValueError):
        return None


def parse_auctions(xlsm_bytes: BytesIO, source_url: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsm_bytes, read_only=True, data_only=True, keep_vba=False)
    ws = wb["Operacje"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        type_op = (d.get("TypOperacji") or "").strip()
        if type_op not in AUCTION_TYPE_OPS:
            continue
        type_tx = (d.get("TypTransakcji") or "").strip()
        if type_tx not in ("S", "O"):
            continue
        isin = d.get("KodISIN")
        if not isinstance(isin, str) or not isin:
            continue
        auction_id = _to_id(d.get("#ID"))
        if auction_id is None:
            continue
        auction_date = _to_date(d.get("DataTransakcji"))
        if auction_date is None:
            continue
        seria = (d.get("Seria") or "").strip()

        out.append({
            "auction_id": auction_id,
            "auction_date": auction_date.isoformat(),
            "settle_date": _to_date(d.get("DataRozliczenia")).isoformat()
                if _to_date(d.get("DataRozliczenia")) else None,
            "type_tx": type_tx,
            "type_op": type_op,
            "seria": seria,
            "isin": isin,
            "maturity_date": _to_date(d.get("DataWykupu")).isoformat()
                if _to_date(d.get("DataWykupu")) else None,
            "years_to_maturity": int(_to_float(d.get("LataDoWykupu")))
                if _to_float(d.get("LataDoWykupu")) is not None else None,
            "coupon_kind": (d.get("Oprocentowanie") or None),
            "offer_min_mln": _to_float(d.get("Podaż Min")
                                       or d.get("PodażMin")),
            "offer_max_mln": _to_float(d.get("Podaż Max")
                                       or d.get("PodażMax")),
            "demand_total_mln": _to_float(d.get("Popyt Łączny")
                                          or d.get("PopytŁączny")),
            "demand_nc_mln": _to_float(d.get("Popyt NK")
                                       or d.get("PopytNK")),
            "sold_total_mln": _to_float(d.get("Sprzedaż lub Odkup Łącznie")
                                        or d.get("SprzedażLubOdkupŁącznie")),
            "sold_nc_mln": _to_float(d.get("Sprzedaż NK")
                                     or d.get("SprzedażNK")),
            "price_min": _to_float(d.get("Cena Min Lub Zamiany")
                                   or d.get("CenaMinLubZamiany")),
            "price_avg": _to_float(d.get("Cena Śr")
                                   or d.get("CenaŚr")),
            "price_max": _to_float(d.get("Cena Max")
                                   or d.get("CenaMax")),
            "yield_max": _yield_to_pct(d.get("Rent Max")
                                       or d.get("RentMax")),
            "yield_avg": _yield_to_pct(d.get("Rent Śr")
                                       or d.get("RentŚr")),
            "yield_min": _yield_to_pct(d.get("Rent Min")
                                       or d.get("RentMin")),
            "source": "mf_xlsm",
            "source_url": source_url,
        })

    return out


def main() -> None:
    print("[1/4] Locating MF Obligacje_Hurtowe.xlsm URL...", flush=True)
    url = find_xlsm_url()
    print(f"  -> {url}", flush=True)

    print("[2/4] Downloading XLSM...", flush=True)
    xlsm = download_xlsm(url)
    print(f"  -> {xlsm.getbuffer().nbytes / 1024:.0f} KB", flush=True)

    print("[3/4] Parsing Operacje (auction-type ops only)...", flush=True)
    rows = parse_auctions(xlsm, url)
    by_op: dict[str, int] = {}
    by_tx: dict[str, int] = {}
    for r in rows:
        by_op[r["type_op"]] = by_op.get(r["type_op"], 0) + 1
        by_tx[r["type_tx"]] = by_tx.get(r["type_tx"], 0) + 1
    print(f"  -> {len(rows)} auction-leg rows", flush=True)
    print(f"  -> by TypOperacji: {dict(sorted(by_op.items(), key=lambda x: -x[1]))}", flush=True)
    print(f"  -> by TypTransakcji: {by_tx}", flush=True)

    # Show most recent 5 sale-side auctions (AS/AU primary interest)
    recent = sorted(
        [r for r in rows if r["type_tx"] == "S" and r["type_op"] in ("AS", "AU")],
        key=lambda r: r["auction_date"],
        reverse=True,
    )[:5]
    if recent:
        print("  -> 5 most recent sale auctions (AS/AU):", flush=True)
        for r in recent:
            bc = (r["demand_total_mln"] or 0) / (r["sold_total_mln"] or 1) \
                if r["sold_total_mln"] else None
            tail_bp = ((r["yield_max"] or 0) - (r["yield_avg"] or 0)) * 100 \
                if r["yield_max"] is not None and r["yield_avg"] is not None else None
            print(
                f"     {r['auction_date']}  {r['seria']:<8}  "
                f"sold={r['sold_total_mln']:>8.1f} mln  "
                f"B/C={bc:.2f}  "
                f"tail={tail_bp:.1f}bp  "
                f"yld_avg={r['yield_avg']:.3f}%" if bc and tail_bp is not None and r['yield_avg']
                else f"     {r['auction_date']}  {r['seria']:<8}  (partial data)",
                flush=True,
            )

    print("[4/4] Upserting to bond_auctions...", flush=True)
    posted = upsert("bond_auctions", rows, on_conflict="auction_id", batch_size=1000)
    print(f"  -> {posted} rows posted", flush=True)
    print("Done.", flush=True)


if __name__ == "__main__":
    main()

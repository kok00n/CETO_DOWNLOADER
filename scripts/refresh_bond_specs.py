"""Refresh bond_specs from the MF 'Obligacje hurtowe' XLSM.

Source: https://www.gov.pl/web/finanse/bony-i-obligacje-hurtowe1
"""

from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

import openpyxl

# allow `python scripts/refresh_bond_specs.py` from repo root
sys.path.insert(0, str(Path(__file__).parent))
from lib.mf_xlsm import download_xlsm, find_xlsm_url  # noqa: E402
from lib.supabase import upsert  # noqa: E402


# Coupon frequency per series prefix.
# Sources: BondSpot TBSP_instruments docs + MF rozporzadzenie.
PREFIX_FREQ: dict[str, int] = {
    # zero-coupon
    "OK": 0, "KO": 0,
    # annual fixed-rate (PFI)
    "PS": 1, "DS": 1, "WS": 1, "OS": 1,
    "AS": 1, "RP": 1, "TK": 1, "CK": 1, "PK": 1, "DK": 1, "PP": 1,
    # annual inflation-linked
    "IZ": 1,
    # semi-annual WIBOR6M floaters
    "WZ": 2, "DZ": 2,
    # POLSTR compound floaters (in arrears) - tentative quarterly
    "NZ": 4,
}


def parse_specs(xlsm_bytes: BytesIO, source_url: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsm_bytes, read_only=True, data_only=True, keep_vba=False)
    ws = wb["Operacje"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    by_isin: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        isin = d.get("KodISIN")
        if not isin or not isinstance(isin, str):
            continue

        seria = (d.get("Seria") or "").strip()
        prefix = seria[:2]
        oprocentowanie = d.get("Oprocentowanie") or "S"
        kupon_raw = d.get("Kupon")
        is_floating = (kupon_raw == "FRN") or (oprocentowanie == "Z")

        coupon_rate = None
        if not is_floating and kupon_raw not in (None, "", "FRN"):
            try:
                coupon_rate = float(kupon_raw)
            except (TypeError, ValueError):
                coupon_rate = None

        maturity = d.get("DataWykupu")
        if maturity is None:
            continue  # skip rows without maturity
        maturity_iso = maturity.date().isoformat() if hasattr(maturity, "date") else str(maturity)[:10]

        transaction = d.get("DataTransakcji")
        transaction_iso = (
            transaction.date().isoformat() if hasattr(transaction, "date") else None
        )

        rec = by_isin.get(isin)
        if rec is None:
            by_isin[isin] = {
                "isin": isin,
                "name": seria,
                "bond_type": prefix,
                "coupon_kind": oprocentowanie,
                "is_floating": is_floating,
                "issue_date": transaction_iso,  # provisional, will MIN below
                "maturity_date": maturity_iso,
                "coupon_rate": coupon_rate,
                "coupon_freq": PREFIX_FREQ.get(prefix),
                "source": "mf_xlsm",
                "source_url": source_url,
            }
        else:
            # Keep earliest transaction date as issue_date proxy
            if transaction_iso and (
                rec["issue_date"] is None or transaction_iso < rec["issue_date"]
            ):
                rec["issue_date"] = transaction_iso

    return list(by_isin.values())


def main() -> None:
    print("[1/4] Locating MF Obligacje_Hurtowe.xlsm URL...", flush=True)
    url = find_xlsm_url()
    print(f"  -> {url}", flush=True)

    print("[2/4] Downloading XLSM...", flush=True)
    xlsm = download_xlsm(url)
    size_kb = xlsm.getbuffer().nbytes / 1024
    print(f"  -> {size_kb:.0f} KB", flush=True)

    print("[3/4] Parsing 'Operacje' sheet...", flush=True)
    specs = parse_specs(xlsm, url)
    by_type: dict[str, int] = {}
    floaters = 0
    for s in specs:
        by_type[s["bond_type"]] = by_type.get(s["bond_type"], 0) + 1
        if s["is_floating"]:
            floaters += 1
    print(f"  -> {len(specs)} unique ISINs ({floaters} floaters)", flush=True)
    print(f"  -> by prefix: {dict(sorted(by_type.items(), key=lambda x: -x[1]))}", flush=True)

    print(f"[4/4] Upserting to bond_specs (batched)...", flush=True)
    posted = upsert("bond_specs", specs, on_conflict="isin")
    print(f"  -> {posted} rows posted", flush=True)
    print("Done.", flush=True)


if __name__ == "__main__":
    main()

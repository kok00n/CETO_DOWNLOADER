"""Refresh bond_outstanding from the MF 'Obligacje hurtowe' XLSM.

Algorithm:
  1. Read Operacje sheet (every wholesale-bond auction/buyback since 1994).
     MF encodes sign in the amount itself: sales positive, buybacks negative.
  2. For each (isin, settlement_date): sum signed amounts.
  3. Per ISIN, sort by settlement date, compute running balance.
  4. Add a synthetic 'redemption' entry on each bond's DataWykupu (maturity)
     with delta = -final_balance, balance = 0 (so post-maturity queries
     correctly return 0 outstanding).
  5. Reconcile against the Zadluzenie sheet (snapshot of currently outstanding
     debt per active ISIN, ground-truth as of a specific date listed in the
     sheet header). Where our forward reconstruction disagrees with the
     snapshot, append a 'reconciliation' entry at the snapshot date with the
     delta needed to match. Historical balances stay reconstructed; current
     snapshot becomes consistent with MF's own truth.
  6. Upsert to bond_outstanding.

Amounts are in PLN mln (zgodnie z konwencja MF).
"""

from __future__ import annotations

import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from lib.mf_xlsm import download_xlsm, find_xlsm_url  # noqa: E402
from lib.supabase import upsert  # noqa: E402


_SNAPSHOT_DATE_RX = re.compile(r"stanu\s+na\s+(\d{2})\.(\d{2})\.(\d{4})", re.IGNORECASE)


def _to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_zadluzenie(xlsm_bytes: BytesIO) -> tuple[dict[str, float], date | None]:
    """Parse the 'Zadluzenie' sheet: snapshot of current outstanding per ISIN.

    Returns (balances_by_isin, snapshot_date).  Snapshot date is parsed from
    a header cell like "Zadluzenie (lacznie) wedlug stanu na 15.05.2026 roku ="
    and falls back to None if not found.
    """
    wb = openpyxl.load_workbook(xlsm_bytes, read_only=True, data_only=True, keep_vba=False)
    ws = wb["Zadłużenie"]

    snapshot_date: date | None = None
    balances: dict[str, float] = {}

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # Try to extract snapshot date from any text cell at the start
        first = row[0]
        if isinstance(first, str) and snapshot_date is None:
            m = _SNAPSHOT_DATE_RX.search(first)
            if m:
                dd, mm, yyyy = m.groups()
                try:
                    snapshot_date = date(int(yyyy), int(mm), int(dd))
                except ValueError:
                    pass

        # Bond rows have shape: Seria, DataWykupu, KodISIN, Kupon, PLN_mln, ...
        if len(row) >= 5 and isinstance(row[2], str) and row[2].startswith("PL"):
            isin = row[2].strip()
            bal = row[4]
            if bal is None:
                continue
            try:
                balances[isin] = float(bal)
            except (TypeError, ValueError):
                continue

    return balances, snapshot_date


def parse_outstanding(xlsm_bytes: BytesIO, source_url: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsm_bytes, read_only=True, data_only=True, keep_vba=False)
    ws = wb["Operacje"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # Step 1: aggregate deltas by (isin, settlement_date).
    # Track tez MIN(auction_date) per (isin, settle_date) - chart 3 dashboard
    # uzywa go zamiast settle_date zeby aukcje pojawialy sie na osi czasu w
    # dniu transakcji a nie 2 dni pozniej. PK pozostaje (isin, change_date)
    # = (isin, settle_date) zeby reconciliation match-owal MF Zadluzenie.
    deltas: dict[str, dict[date, float]] = defaultdict(lambda: defaultdict(float))
    op_kinds: dict[str, dict[date, set[str]]] = defaultdict(lambda: defaultdict(set))
    auction_dates: dict[str, dict[date, date]] = defaultdict(dict)
    maturity_by_isin: dict[str, date] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        isin = d.get("KodISIN")
        if not isinstance(isin, str) or not isin:
            continue
        settle = _to_date(d.get("DataRozliczenia")) or _to_date(d.get("DataTransakcji"))
        if settle is None:
            continue
        # MF encodes sign in the amount itself: sales positive, buybacks
        # negative. We sum as-is and only use TypTransakcji to label op_type.
        amount = _to_float(d.get("SprzedażLubOdkupŁącznie")) or _to_float(
            d.get("Sprzedaż lub Odkup Łącznie")
        )
        if amount is None or amount == 0:
            continue
        type_tx = (d.get("TypTransakcji") or "").strip()
        if type_tx not in ("S", "O"):
            continue
        deltas[isin][settle] += amount
        op_kinds[isin][settle].add("sale" if type_tx == "S" else "buyback")

        auction = _to_date(d.get("DataTransakcji"))
        if auction is not None:
            prior = auction_dates[isin].get(settle)
            if prior is None or auction < prior:
                auction_dates[isin][settle] = auction

        mat = _to_date(d.get("DataWykupu"))
        if mat:
            maturity_by_isin[isin] = mat

    # Step 2: parse Zadluzenie snapshot for reconciliation truth
    xlsm_bytes.seek(0)
    zad_balances, snapshot_date = parse_zadluzenie(xlsm_bytes)
    if snapshot_date is None:
        print("  ! WARN: snapshot date not found in Zadluzenie header - "
              "reconciliation will be skipped", flush=True)

    # Step 3: per ISIN, sort dates, compute running balance, then reconcile
    #         against Zadluzenie, then add redemption at maturity
    rows: list[dict] = []
    for isin, by_date in deltas.items():
        sorted_dates = sorted(by_date.keys())
        balance = 0.0
        balance_at_snapshot: float | None = None
        for cd in sorted_dates:
            delta = by_date[cd]
            balance += delta
            kinds = op_kinds[isin][cd]
            op_type = next(iter(kinds)) if len(kinds) == 1 else "mixed"
            ad = auction_dates[isin].get(cd)
            rows.append({
                "isin": isin,
                "change_date": cd.isoformat(),
                "auction_date": ad.isoformat() if ad else None,
                "delta_mln_pln": round(delta, 3),
                "balance_mln_pln": round(max(balance, 0.0), 3),
                "op_type": op_type,
                "source_url": source_url,
            })
            if snapshot_date and cd <= snapshot_date:
                balance_at_snapshot = balance

        last_change_date = sorted_dates[-1] if sorted_dates else None
        mat = maturity_by_isin.get(isin)

        # Step 3a: reconciliation entry at snapshot_date.
        # Two cases reconcile:
        #   - bond is in Zadluzenie -> use zad balance as target
        #   - bond is NOT in Zadluzenie but our calc says active and maturity
        #     is in future -> bond was extinguished early; target = 0
        if snapshot_date and last_change_date and snapshot_date >= sorted_dates[0]:
            target: float | None = zad_balances.get(isin)
            if target is None and balance > 0.001 and mat and mat > snapshot_date:
                target = 0.0
            if target is not None and balance_at_snapshot is not None:
                diff = target - balance_at_snapshot
                if abs(diff) > 0.005:
                    recon_date = snapshot_date
                    # avoid PK clash with real op on same date
                    if recon_date in by_date:
                        recon_date = recon_date + timedelta(days=1)
                    rows.append({
                        "isin": isin,
                        "change_date": recon_date.isoformat(),
                        "auction_date": None,  # recon nie z aukcji
                        "delta_mln_pln": round(diff, 3),
                        "balance_mln_pln": round(target, 3),
                        "op_type": "reconciliation",
                        "source_url": source_url,
                    })
                    balance = target  # update for redemption check below

        # Step 3b: synthetic redemption row at maturity if final balance
        # still > 0. If maturity coincides with the last recorded op
        # (typical for zero-coupon bonds: last KZ op settled on maturity
        # day), push redemption to mat+1 - bond is alive on its maturity
        # day, redeemed EOD.
        if mat and balance > 0.001:
            last_rec = max(
                (date.fromisoformat(r["change_date"]) for r in rows if r["isin"] == isin),
                default=None,
            )
            if last_rec is None or mat > last_rec:
                redemption_date = mat
            elif mat == last_rec:
                redemption_date = mat + timedelta(days=1)
            else:
                redemption_date = None  # maturity before last op (shouldn't happen)
            if redemption_date is not None:
                rows.append({
                    "isin": isin,
                    "change_date": redemption_date.isoformat(),
                    "auction_date": None,  # redemption nie z aukcji
                    "delta_mln_pln": -round(balance, 3),
                    "balance_mln_pln": 0.0,
                    "op_type": "redemption",
                    "source_url": source_url,
                })

    return rows


def main() -> None:
    print("[1/4] Locating MF Obligacje_Hurtowe.xlsm URL...", flush=True)
    url = find_xlsm_url()
    print(f"  -> {url}", flush=True)

    print("[2/4] Downloading XLSM...", flush=True)
    xlsm = download_xlsm(url)
    print(f"  -> {xlsm.getbuffer().nbytes / 1024:.0f} KB", flush=True)

    print("[3/4] Parsing Operacje sheet, building running balances...", flush=True)
    rows = parse_outstanding(xlsm, url)
    distinct_isins = len({r["isin"] for r in rows})
    redemptions = sum(1 for r in rows if r["op_type"] == "redemption")
    print(
        f"  -> {len(rows)} balance-change rows across {distinct_isins} ISINs "
        f"({redemptions} synthetic redemptions)",
        flush=True,
    )

    # Print a few sanity samples (top 5 ISINs by current balance)
    latest_by_isin: dict[str, dict] = {}
    for r in rows:
        prev = latest_by_isin.get(r["isin"])
        if prev is None or r["change_date"] > prev["change_date"]:
            latest_by_isin[r["isin"]] = r
    active = [r for r in latest_by_isin.values() if r["balance_mln_pln"] > 0]
    active.sort(key=lambda r: r["balance_mln_pln"], reverse=True)
    print("  -> top 5 by current balance (PLN mln):", flush=True)
    for r in active[:5]:
        print(f"     {r['isin']}  {r['balance_mln_pln']:>14,.1f}  "
              f"(last change {r['change_date']}, {r['op_type']})", flush=True)

    print("[4/4] Upserting to bond_outstanding (batched)...", flush=True)
    posted = upsert("bond_outstanding", rows, on_conflict="isin,change_date", batch_size=1000)
    print(f"  -> {posted} rows posted", flush=True)
    print("Done.", flush=True)


if __name__ == "__main__":
    main()
